"""
Report Generator - Creates Excel, CSV, and JSON reports from analysis results
"""
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

from app.belt_analyzer import AnalysisResult, SegmentMeasurement

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate reports in various formats from belt analysis results"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_all(self, result: AnalysisResult, base_name: Optional[str] = None) -> dict:
        """Generate reports in all formats"""
        if base_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"belt_analysis_{timestamp}"
        
        paths = {
            "excel": self.generate_excel(result, f"{base_name}.xlsx"),
            "csv": self.generate_csv(result, f"{base_name}.csv"),
            "json": self.generate_json(result, f"{base_name}.json")
        }
        
        logger.info(f"Generated reports: {paths}")
        return paths
    
    def generate_excel(self, result: AnalysisResult, filename: str) -> str:
        """Generate detailed Excel report with formatting and charts"""
        filepath = self.output_dir / filename
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        alert_fill = PatternFill("solid", fgColor="FFC7CE")
        good_fill = PatternFill("solid", fgColor="C6EFCE")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Podsumowanie"
        
        summary_data = [
            ["Raport Analizy Taśmy Przenośnika", ""],
            ["", ""],
            ["Plik źródłowy:", result.source_file],
            ["Data analizy:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Całkowita liczba klatek:", result.total_frames],
            ["FPS:", result.fps],
            ["Liczba wykrytych segmentów:", len(result.segments)],
            ["Liczba alertów:", len(result.alerts)],
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif col_idx == 1:
                    cell.font = Font(bold=True)
        
        # Segments Sheet
        ws_segments = wb.create_sheet("Segmenty")
        
        headers = [
            "ID Segmentu", "Klatka Start", "Klatka Koniec",
            "Min. Szerokość (px)", "Max. Szerokość (px)", 
            "Śr. Szerokość (px)", "Wariancja", "Liczba Pomiarów", "Status"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_segments.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Calculate average width across all segments for comparison
        all_widths = [w for s in result.segments for w in s.widths]
        global_avg = sum(all_widths) / len(all_widths) if all_widths else 0
        
        for row_idx, segment in enumerate(result.segments, 2):
            # Determine status based on width variance
            status = "OK"
            fill = good_fill
            if segment.width_variance > 100:
                status = "UWAGA"
                fill = alert_fill
            elif segment.min_width < global_avg * 0.9:
                status = "UWAGA"
                fill = alert_fill
            
            row_data = [
                segment.segment_id,
                segment.frame_start,
                segment.frame_end,
                round(segment.min_width, 2),
                round(segment.max_width, 2),
                round(segment.avg_width, 2),
                round(segment.width_variance, 2),
                len(segment.widths),
                status
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_segments.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if col_idx == 9:  # Status column
                    cell.fill = fill
        
        # Auto-width columns
        for col_idx, _ in enumerate(headers, 1):
            ws_segments.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Alerts Sheet
        ws_alerts = wb.create_sheet("Alerty")
        alert_headers = ["Typ", "Klatka", "Wiadomość", "Poziom"]
        
        for col_idx, header in enumerate(alert_headers, 1):
            cell = ws_alerts.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, alert in enumerate(result.alerts, 2):
            row_data = [
                alert.get("type", ""),
                alert.get("frame", ""),
                alert.get("message", ""),
                alert.get("severity", "")
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_alerts.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if alert.get("severity") == "warning":
                    cell.fill = alert_fill
        
        # Width Chart (if enough data)
        if len(result.segments) >= 2:
            ws_chart = wb.create_sheet("Wykres")
            ws_chart['A1'] = "Segment"
            ws_chart['B1'] = "Min"
            ws_chart['C1'] = "Max"
            ws_chart['D1'] = "Średnia"
            
            for i, seg in enumerate(result.segments, 2):
                ws_chart[f'A{i}'] = seg.segment_id
                ws_chart[f'B{i}'] = seg.min_width
                ws_chart[f'C{i}'] = seg.max_width
                ws_chart[f'D{i}'] = seg.avg_width
            
            chart = LineChart()
            chart.title = "Szerokość Taśmy per Segment"
            chart.y_axis.title = "Szerokość (px)"
            chart.x_axis.title = "Segment"
            
            data = Reference(ws_chart, min_col=2, max_col=4, min_row=1, max_row=len(result.segments)+1)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(result.segments)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_chart.add_chart(chart, "F2")
        
        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")
        return str(filepath)
    
    def generate_csv(self, result: AnalysisResult, filename: str) -> str:
        """Generate CSV report"""
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow([
                "segment_id", "frame_start", "frame_end",
                "min_width_px", "max_width_px", "avg_width_px",
                "variance", "measurement_count"
            ])
            
            # Data
            for segment in result.segments:
                writer.writerow([
                    segment.segment_id,
                    segment.frame_start,
                    segment.frame_end,
                    round(segment.min_width, 2),
                    round(segment.max_width, 2),
                    round(segment.avg_width, 2),
                    round(segment.width_variance, 2),
                    len(segment.widths)
                ])
        
        logger.info(f"CSV report saved: {filepath}")
        return str(filepath)
    
    def generate_json(self, result: AnalysisResult, filename: str) -> str:
        """Generate JSON report"""
        filepath = self.output_dir / filename
        
        data = result.to_dict()
        data["generated_at"] = datetime.now().isoformat()
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON report saved: {filepath}")
        return str(filepath)
